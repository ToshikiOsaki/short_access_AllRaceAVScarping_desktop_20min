import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
import sys
import threading
from tkinter import messagebox
import requests
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl.utils import get_column_letter
import os
import json
import urllib.parse
import time
import random

# ログをGUIに出力するためのクラス（標準出力を乗っ取る）
class TextRedirector(object):
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        # スレッドセーフにGUIを更新
        self.widget.after(0, self._write, str)

    def _write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.see("end")
        self.widget.configure(state="disabled")
    
    def flush(self):
        pass

class NetkeibaScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Netkeiba 平均タイム取得アプリ アクセス短縮版")
        self.root.geometry("1100x900")
        
        # 背景色を薄ピンクに設定
        bg_color = "#FFE4E1" # MistyRose
        self.root.configure(bg=bg_color)

        # スタイルの設定
        style = ttk.Style()
        style.theme_use('default')
        style.configure('.', background=bg_color)
        style.configure('TLabel', background=bg_color)
        style.configure('TFrame', background=bg_color)
        style.configure('TCheckbutton', background=bg_color)
        style.configure('TLabelframe', background=bg_color)
        style.configure('TLabelframe.Label', background=bg_color)
        style.map('TCheckbutton', background=[('active', bg_color)])

        # メインフレーム
        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- 設定エリア（スクロール可能にする） ---
        settings_canvas = tk.Canvas(main_frame, bg=bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=settings_canvas.yview)
        self.scrollable_frame = ttk.Frame(settings_canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: settings_canvas.configure(
                scrollregion=settings_canvas.bbox("all")
            )
        )

        settings_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        settings_canvas.configure(yscrollcommand=scrollbar.set)

        settings_canvas.pack(side="top", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 1. 芝・ダート (siba_dart_number)
        self.create_checkbox_group("トラック (siba_dart_number)", 
                                   {"芝": "1", "ダート": "2"}, 
                                   "track_vars", 0, 0)

        # 2. 競馬場 (keibajyou_number)
        self.create_checkbox_group("競馬場 (keibajyou_number)", 
                                   {"札幌": "01", "函館": "02", "福島": "03", "新潟": "04", 
                                    "東京": "05", "中山": "06", "中京": "07", "京都": "08", 
                                    "阪神": "09", "小倉": "10"}, 
                                   "jyo_vars", 0, 1)

        # 3. 馬場状態 (babajyoutai)
        self.create_checkbox_group("馬場状態 (babajyoutai)", 
                                   {"良": "1", "稍重": "2", "重": "3", "不良": "4"}, 
                                   "baba_vars", 1, 0)

        # 4. クラス (grade_sentence)
        self.create_checkbox_group("クラス (grade_sentence)", 
                                   {"重賞・OP": "&grade%5B%5D=1&grade%5B%5D=2&grade%5B%5D=3&grade%5B%5D=11&grade%5B%5D=4",
                                    "3勝クラス": "&grade%5B%5D=5",
                                    "2勝クラス": "&grade%5B%5D=6",
                                    "1勝クラス": "&grade%5B%5D=7",
                                    "未勝利": "&grade%5B%5D=9",
                                    "新馬": "&grade%5B%5D=8"}, 
                                   "grade_vars", 1, 1)

        # 5. 距離 (kyori)
        kyori_list = ["1000以下","1100","1150","1200","1300","1400","1500","1600","1700","1800","1900","2000","2100","2200","2300","2400",
                      "2500","2600","3000","3200","3400","3600"]
        self.create_checkbox_group("距離 (kyori)", 
                                   {k: k for k in kyori_list}, 
                                   "kyori_vars", 2, 0, colspan=2)

        # 6. 期間とファイルパス
        input_frame = ttk.LabelFrame(self.scrollable_frame, text="設定", padding="10")
        input_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

        ttk.Label(input_frame, text="開始年:").grid(row=0, column=0, padx=5)
        self.start_year_var = tk.StringVar(value="2020")
        ttk.Entry(input_frame, textvariable=self.start_year_var, width=10).grid(row=0, column=1, padx=5)

        ttk.Label(input_frame, text="終了年:").grid(row=0, column=2, padx=5)
        self.end_year_var = tk.StringVar(value="2025")
        ttk.Entry(input_frame, textvariable=self.end_year_var, width=10).grid(row=0, column=3, padx=5)

        ttk.Label(input_frame, text="開始月:").grid(row=0, column=4, padx=5)
        self.start_month_var = tk.StringVar(value="1")
        month_values = [str(i) for i in range(1, 13)]
        ttk.Combobox(input_frame, textvariable=self.start_month_var, values=month_values, width=3, state="readonly").grid(row=0, column=5, padx=5)

        ttk.Label(input_frame, text="終了月:").grid(row=0, column=6, padx=5)
        self.end_month_var = tk.StringVar(value="12")
        ttk.Combobox(input_frame, textvariable=self.end_month_var, values=month_values, width=3, state="readonly").grid(row=0, column=7, padx=5)

        # 集計期間とスライド間隔の設定を追加
        ttk.Label(input_frame, text="集計期間(年):").grid(row=1, column=0, padx=5, pady=5)
        self.agg_years_var = tk.StringVar(value="1")
        ttk.Entry(input_frame, textvariable=self.agg_years_var, width=10).grid(row=1, column=1, padx=5)

        ttk.Label(input_frame, text="スライド間隔(年):").grid(row=1, column=2, padx=5, pady=5)
        self.slide_years_var = tk.StringVar(value="1")
        ttk.Entry(input_frame, textvariable=self.slide_years_var, width=10).grid(row=1, column=3, padx=5)

        ttk.Label(input_frame, text="保存先ファイルパス:").grid(row=2, column=0, padx=5, pady=5)
        default_path = os.path.join(os.path.expanduser("~"), "Documents", "netkeiba_average_time.xlsx")
        self.file_path_var = tk.StringVar(value=default_path)
        ttk.Entry(input_frame, textvariable=self.file_path_var, width=80).grid(row=2, column=1, columnspan=7, padx=5, pady=5)

        # --- ステータス表示 ---
        self.status_var = tk.StringVar(value="")
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var, font=("Meiryo", 12, "bold"))
        self.status_label.pack(pady=5)

        # --- 実行・中断ボタン ---
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        self.run_button = ttk.Button(button_frame, text="スクレイピング開始", command=self.start_scraping)
        self.run_button.pack(side=tk.LEFT, padx=5)
        self.stop_button = ttk.Button(button_frame, text="中断", command=self.stop_scraping, state="disabled")
        self.stop_button.pack(side=tk.LEFT, padx=5)

        self.is_running = False

        # --- ログ表示エリア ---
        log_frame = ttk.LabelFrame(main_frame, text="ターミナル出力", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, state='disabled', height=15)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        sys.stdout = TextRedirector(self.log_text)
        sys.stderr = TextRedirector(self.log_text, "stderr")
        self.log_text.tag_config("stderr", foreground="red")

        # --- 設定の読み込みと終了処理 ---
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_name = os.path.splitext(os.path.basename(__file__))[0]
        self.config_file = os.path.join(base_dir, f"{file_name}_config.json")
        self.load_config()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_checkbox_group(self, title, options, var_name, row, col, colspan=1):
        frame = ttk.LabelFrame(self.scrollable_frame, text=title, padding="5")
        frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=5, pady=5)
        
        vars_dict = {}
        r, c = 0, 0
        max_col = 6

        for label, value in options.items():
            var = tk.BooleanVar(value=True)
            chk = ttk.Checkbutton(frame, text=label, variable=var)
            chk.grid(row=r, column=c, sticky="w", padx=2)
            vars_dict[value] = var
            
            c += 1
            if c >= max_col:
                c = 0
                r += 1
        
        setattr(self, var_name, vars_dict)

    def get_selected_values(self, var_name):
        vars_dict = getattr(self, var_name)
        return [val for val, var in vars_dict.items() if var.get()]

    def start_scraping(self):
        try:
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
            agg_years = int(self.agg_years_var.get())
            slide_years = int(self.slide_years_var.get())

            test_periods = []
            current_start = start_year
            while True:
                current_end = current_start + agg_years - 1
                if current_end > end_year:
                    break
                test_periods.append((current_start, current_end))
                current_start += slide_years
            
            if not test_periods:
                messagebox.showerror("設定エラー", "指定された条件では取得可能な期間がありません。\n集計期間が指定範囲より長い可能性があります。")
                return

            last_end = test_periods[-1][1]
            if last_end < end_year:
                msg = f"指定された終了年は {end_year} 年ですが、\n設定に基づくと最後のデータ取得は {last_end} 年までとなります。\n\n({last_end + 1}年～{end_year}年のデータは集計期間不足のため含まれません)\n\n処理を開始してもよろしいですか？"
                if not messagebox.askyesno("期間確認", msg):
                    return
        except ValueError:
            messagebox.showerror("入力エラー", "年数設定には半角数字を入力してください。")
            return
        
        self.status_var.set("データ取得中")
        self.status_label.configure(foreground="blue")
        
        self.run_button.config(state="disabled")
        self.stop_button.config(state="normal")
        self.is_running = True
        
        thread = threading.Thread(target=self.run_scraping_logic)
        thread.daemon = True
        thread.start()

    def save_workbook(self, wb, file_path):
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        wb.save(file_path)

    def stop_scraping(self):
        if self.is_running:
            self.is_running = False
            self.status_var.set("中断処理中...")
            print("中断ボタンが押されました。現在の処理が完了次第停止します。")

    def on_closing(self):
        self.save_config()
        self.root.destroy()

    def save_config(self):
        save_path = self.file_path_var.get()
        home_dir = os.path.expanduser("~")
        if save_path.startswith(home_dir):
            save_path = save_path.replace(home_dir, "~", 1)

        config = {
            "start_year": self.start_year_var.get(),
            "end_year": self.end_year_var.get(),
            "start_month": self.start_month_var.get(),
            "end_month": self.end_month_var.get(),
            "agg_years": self.agg_years_var.get(),
            "slide_years": self.slide_years_var.get(),
            "file_path": save_path,
            "checkboxes": {}
        }
        for var_name in ["track_vars", "jyo_vars", "baba_vars", "grade_vars", "kyori_vars"]:
            vars_dict = getattr(self, var_name)
            config["checkboxes"][var_name] = {key: var.get() for key, var in vars_dict.items()}

        try:
            with open(self.config_file, 'w') as f:
                json.dump(config, f, indent=4)
            print("設定を保存しました。")
        except Exception as e:
            print(f"設定の保存中にエラーが発生しました: {e}")

    def load_config(self):
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    config = json.load(f)

                self.start_year_var.set(config.get("start_year", "2020"))
                self.end_year_var.set(config.get("end_year", "2025"))
                self.start_month_var.set(config.get("start_month", "1"))
                self.end_month_var.set(config.get("end_month", "12"))
                self.agg_years_var.set(config.get("agg_years", "1"))
                self.slide_years_var.set(config.get("slide_years", "1"))

                loaded_path = config.get("file_path", "")
                if loaded_path.startswith("~"):
                    loaded_path = os.path.expanduser(loaded_path)

                default_path = os.path.join(os.path.expanduser("~"), "Documents", "netkeiba_average_time.xlsx")

                if loaded_path and os.path.exists(os.path.dirname(loaded_path)):
                    self.file_path_var.set(loaded_path)
                else:
                    self.file_path_var.set(default_path)

                if "checkboxes" in config:
                    for var_name, states in config["checkboxes"].items():
                        vars_dict = getattr(self, var_name)
                        for key, state in states.items():
                            if key in vars_dict:
                                vars_dict[key].set(state)
                print("前回の設定を読み込みました。")
        except Exception as e:
            print(f"設定の読み込み中にエラーが発生しました: {e}")

    def align_excel_rows(self, input_path, output_path):
        print(f"\n--- データの整形処理を開始します ---")
        try:
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active
        except FileNotFoundError:
            print(f"エラー: ファイルが見つかりません: {input_path}")
            return

        all_conditions_list = []
        processed_blocks = []

        max_col = ws.max_column
        for col_idx in range(1, max_col + 1, 3):
            year_header = ws.cell(row=1, column=col_idx).value
            if not year_header or " 条件" not in year_header:
                continue
            
            headers = [ws.cell(row=1, column=col_idx + i).value for i in range(3)]
            year = year_header.replace(" 条件", "")
            
            year_data = {}
            current_year_conditions = []
            for row_idx in range(2, ws.max_row + 1):
                condition = ws.cell(row=row_idx, column=col_idx).value
                if condition and str(condition).strip():
                    sample_count = ws.cell(row=row_idx, column=col_idx + 1).value
                    avg_time = ws.cell(row=row_idx, column=col_idx + 2).value
                    year_data[condition] = (sample_count, avg_time)
                    current_year_conditions.append(condition)
                    
            processed_blocks.append({
                'year': year,
                'headers': headers,
                'data': year_data
            })

            last_index = -1
            for condition in current_year_conditions:
                if condition in all_conditions_list:
                    last_index = all_conditions_list.index(condition)
                else:
                    all_conditions_list.insert(last_index + 1, condition)
                    last_index += 1

        if not all_conditions_list:
            print("処理対象のデータが見つかりませんでした。")
            return

        master_conditions = all_conditions_list
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "整形済みデータ"

        header_row = []
        for block in processed_blocks:
            header_row.extend(block['headers'])
        new_ws.append(header_row)

        for condition in master_conditions:
            new_row_data = []
            for block in processed_blocks:
                year_data = block['data']
                if condition in year_data:
                    sample_count, avg_time = year_data[condition]
                    new_row_data.extend([condition, sample_count, avg_time])
                else:
                    new_row_data.extend([condition, None, None])
            new_ws.append(new_row_data)
        
        # 列幅を自動調整する
        for col in new_ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                char_count = 0
                if cell.value:
                    for char in str(cell.value):
                        if ord(char) <= 255:
                            char_count += 1
                        else:
                            char_count += 2
                if char_count > max_length:
                    max_length = char_count
            new_ws.column_dimensions[column].width = max_length + 2

        try:
            new_wb.save(output_path)
            print(f"整形が完了しました。ファイルを保存しました: {output_path}")
            self.root.after(0, lambda: self.status_var.set("データ取得＆整形完了"))
        except Exception as e:
            print(f"ファイルの保存中にエラーが発生しました: {e}")

    def run_scraping_logic(self):
        start_time = time.time()
        processed_count = 0
        # セッションオブジェクトを作成し、接続を再利用して高速化
        session = requests.Session()
        try:
            # --- GUIから値を取得 ---
            siba_dart_number = self.get_selected_values("track_vars")
            keibajyou_number = self.get_selected_values("jyo_vars")
            grade_sentence = self.get_selected_values("grade_vars")
            kyori = self.get_selected_values("kyori_vars")
            
            # 馬場状態の生成
            baba_values = self.get_selected_values("baba_vars")
            baba_values.sort()
            babajyoutai = ""
            for val in baba_values:
                babajyoutai += f"&baba%5B%5D={val}"
            
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
            start_month = self.start_month_var.get()
            end_month = self.end_month_var.get()
            
            # 集計期間とスライド間隔を取得
            agg_years = int(self.agg_years_var.get())
            slide_years = int(self.slide_years_var.get())

            target_periods = []
            current_start = start_year
            while True:
                current_end = current_start + agg_years - 1
                if current_end > end_year:
                    break
                target_periods.append((current_start, current_end))
                current_start += slide_years
            
            file_path = self.file_path_var.get().strip('"')

            # --- 高速化のためのレース条件リスト ---
            shutoku_jyouken = ["芝札幌新馬1000以下m","芝札幌G1、G2、G3、L、OP1200m","芝札幌3勝(1600万)1200m","芝札幌2勝(1000万)1200m","芝札幌1勝(500万)1200m","芝札幌未勝利1200m","芝札幌新馬1200m","芝札幌G1、G2、G3、L、OP1500m","芝札幌3勝(1600万)1500m","芝札幌2勝(1000万)1500m","芝札幌1勝(500万)1500m","芝札幌未勝利1500m","芝札幌新馬1500m","芝札幌G1、G2、G3、L、OP1800m","芝札幌3勝(1600万)1800m","芝札幌2勝(1000万)1800m","芝札幌1勝(500万)1800m","芝札幌未勝利1800m","芝札幌新馬1800m","芝札幌G1、G2、G3、L、OP2000m","芝札幌3勝(1600万)2000m","芝札幌2勝(1000万)2000m","芝札幌1勝(500万)2000m","芝札幌未勝利2000m","芝札幌新馬2000m","芝札幌G1、G2、G3、L、OP2600m","芝札幌2勝(1000万)2600m","芝札幌1勝(500万)2600m","芝札幌未勝利2600m","芝函館新馬1000以下m","芝函館G1、G2、G3、L、OP1200m","芝函館3勝(1600万)1200m","芝函館2勝(1000万)1200m","芝函館1勝(500万)1200m","芝函館未勝利1200m","芝函館新馬1200m","芝函館G1、G2、G3、L、OP1800m","芝函館3勝(1600万)1800m","芝函館2勝(1000万)1800m","芝函館1勝(500万)1800m","芝函館未勝利1800m","芝函館新馬1800m","芝函館G1、G2、G3、L、OP2000m","芝函館3勝(1600万)2000m","芝函館2勝(1000万)2000m","芝函館1勝(500万)2000m","芝函館未勝利2000m","芝函館新馬2000m","芝函館G1、G2、G3、L、OP2600m","芝函館2勝(1000万)2600m","芝函館1勝(500万)2600m","芝函館未勝利2600m","芝福島G1、G2、G3、L、OP1200m","芝福島3勝(1600万)1200m","芝福島2勝(1000万)1200m","芝福島1勝(500万)1200m","芝福島未勝利1200m","芝福島新馬1200m","芝福島G1、G2、G3、L、OP1800m","芝福島3勝(1600万)1800m","芝福島2勝(1000万)1800m","芝福島1勝(500万)1800m","芝福島未勝利1800m","芝福島新馬1800m","芝福島G1、G2、G3、L、OP2000m","芝福島3勝(1600万)2000m","芝福島2勝(1000万)2000m","芝福島1勝(500万)2000m","芝福島未勝利2000m","芝福島新馬2000m","芝福島2勝(1000万)2600m","芝福島1勝(500万)2600m","芝福島未勝利2600m","芝新潟G1、G2、G3、L、OP1000以下m","芝新潟3勝(1600万)1000以下m","芝新潟2勝(1000万)1000以下m","芝新潟1勝(500万)1000以下m","芝新潟未勝利1000以下m","芝新潟3勝(1600万)1200m","芝新潟2勝(1000万)1200m","芝新潟1勝(500万)1200m","芝新潟未勝利1200m","芝新潟新馬1200m","芝新潟G1、G2、G3、L、OP1400m","芝新潟3勝(1600万)1400m","芝新潟2勝(1000万)1400m","芝新潟1勝(500万)1400m","芝新潟未勝利1400m","芝新潟新馬1400m","芝新潟G1、G2、G3、L、OP1600m","芝新潟3勝(1600万)1600m","芝新潟2勝(1000万)1600m","芝新潟1勝(500万)1600m","芝新潟未勝利1600m","芝新潟新馬1600m","芝新潟G1、G2、G3、L、OP1800m","芝新潟3勝(1600万)1800m","芝新潟2勝(1000万)1800m","芝新潟1勝(500万)1800m","芝新潟未勝利1800m","芝新潟新馬1800m","芝新潟G1、G2、G3、L、OP2000m","芝新潟3勝(1600万)2000m","芝新潟2勝(1000万)2000m","芝新潟1勝(500万)2000m","芝新潟未勝利2000m","芝新潟新馬2000m","芝新潟G1、G2、G3、L、OP2200m","芝新潟3勝(1600万)2200m","芝新潟2勝(1000万)2200m","芝新潟1勝(500万)2200m","芝新潟未勝利2200m","芝新潟2勝(1000万)2400m","芝新潟1勝(500万)2400m","芝新潟未勝利2400m","芝東京G1、G2、G3、L、OP1400m","芝東京3勝(1600万)1400m","芝東京2勝(1000万)1400m","芝東京1勝(500万)1400m","芝東京未勝利1400m","芝東京新馬1400m","芝東京G1、G2、G3、L、OP1600m","芝東京3勝(1600万)1600m","芝東京2勝(1000万)1600m","芝東京1勝(500万)1600m","芝東京未勝利1600m","芝東京新馬1600m","芝東京G1、G2、G3、L、OP1800m","芝東京3勝(1600万)1800m","芝東京2勝(1000万)1800m","芝東京1勝(500万)1800m","芝東京未勝利1800m","芝東京新馬1800m","芝東京G1、G2、G3、L、OP2000m","芝東京3勝(1600万)2000m","芝東京2勝(1000万)2000m","芝東京1勝(500万)2000m","芝東京未勝利2000m","芝東京新馬2000m","芝東京1勝(500万)2300m","芝東京未勝利2300m","芝東京G1、G2、G3、L、OP2400m","芝東京3勝(1600万)2400m","芝東京2勝(1000万)2400m","芝東京1勝(500万)2400m","芝東京未勝利2400m","芝東京G1、G2、G3、L、OP2500m","芝東京G1、G2、G3、L、OP3400m","芝中山G1、G2、G3、L、OP1200m","芝中山3勝(1600万)1200m","芝中山2勝(1000万)1200m","芝中山1勝(500万)1200m","芝中山未勝利1200m","芝中山新馬1200m","芝中山G1、G2、G3、L、OP1600m","芝中山3勝(1600万)1600m","芝中山2勝(1000万)1600m","芝中山1勝(500万)1600m","芝中山未勝利1600m","芝中山新馬1600m","芝中山G1、G2、G3、L、OP1800m","芝中山3勝(1600万)1800m","芝中山2勝(1000万)1800m","芝中山1勝(500万)1800m","芝中山未勝利1800m","芝中山新馬1800m","芝中山G1、G2、G3、L、OP2000m","芝中山3勝(1600万)2000m","芝中山2勝(1000万)2000m","芝中山1勝(500万)2000m","芝中山未勝利2000m","芝中山新馬2000m","芝中山G1、G2、G3、L、OP2200m","芝中山3勝(1600万)2200m","芝中山2勝(1000万)2200m","芝中山1勝(500万)2200m","芝中山未勝利2200m","芝中山G1、G2、G3、L、OP2500m","芝中山3勝(1600万)2500m","芝中山2勝(1000万)2500m","芝中山1勝(500万)2500m","芝中山G1、G2、G3、L、OP3600m","芝中京G1、G2、G3、L、OP1200m","芝中京3勝(1600万)1200m","芝中京2勝(1000万)1200m","芝中京1勝(500万)1200m","芝中京未勝利1200m","芝中京新馬1200m","芝中京G1、G2、G3、L、OP1400m","芝中京3勝(1600万)1400m","芝中京2勝(1000万)1400m","芝中京1勝(500万)1400m","芝中京未勝利1400m","芝中京新馬1400m","芝中京G1、G2、G3、L、OP1600m","芝中京3勝(1600万)1600m","芝中京2勝(1000万)1600m","芝中京1勝(500万)1600m","芝中京未勝利1600m","芝中京新馬1600m","芝中京G1、G2、G3、L、OP2000m","芝中京3勝(1600万)2000m","芝中京2勝(1000万)2000m","芝中京1勝(500万)2000m","芝中京未勝利2000m","芝中京新馬2000m","芝中京G1、G2、G3、L、OP2200m","芝中京3勝(1600万)2200m","芝中京2勝(1000万)2200m","芝中京1勝(500万)2200m","芝中京未勝利2200m","芝中京G1、G2、G3、L、OP3000m","芝京都G1、G2、G3、L、OP1200m","芝京都3勝(1600万)1200m","芝京都2勝(1000万)1200m","芝京都1勝(500万)1200m","芝京都未勝利1200m","芝京都新馬1200m","芝京都G1、G2、G3、L、OP1400m","芝京都3勝(1600万)1400m","芝京都2勝(1000万)1400m","芝京都1勝(500万)1400m","芝京都未勝利1400m","芝京都新馬1400m","芝京都G1、G2、G3、L、OP1600m","芝京都3勝(1600万)1600m","芝京都2勝(1000万)1600m","芝京都1勝(500万)1600m","芝京都未勝利1600m","芝京都新馬1600m","芝京都G1、G2、G3、L、OP1800m","芝京都3勝(1600万)1800m","芝京都2勝(1000万)1800m","芝京都1勝(500万)1800m","芝京都未勝利1800m","芝京都新馬1800m","芝京都G1、G2、G3、L、OP2000m","芝京都3勝(1600万)2000m","芝京都2勝(1000万)2000m","芝京都1勝(500万)2000m","芝京都未勝利2000m","芝京都新馬2000m","芝京都G1、G2、G3、L、OP2200m","芝京都3勝(1600万)2200m","芝京都2勝(1000万)2200m","芝京都1勝(500万)2200m","芝京都未勝利2200m","芝京都G1、G2、G3、L、OP2400m","芝京都3勝(1600万)2400m","芝京都2勝(1000万)2400m","芝京都1勝(500万)2400m","芝京都未勝利2400m","芝京都G1、G2、G3、L、OP3000m","芝京都G1、G2、G3、L、OP3200m","芝阪神G1、G2、G3、L、OP1200m","芝阪神3勝(1600万)1200m","芝阪神2勝(1000万)1200m","芝阪神1勝(500万)1200m","芝阪神未勝利1200m","芝阪神新馬1200m","芝阪神G1、G2、G3、L、OP1400m","芝阪神3勝(1600万)1400m","芝阪神2勝(1000万)1400m","芝阪神1勝(500万)1400m","芝阪神未勝利1400m","芝阪神新馬1400m","芝阪神G1、G2、G3、L、OP1600m","芝阪神3勝(1600万)1600m","芝阪神2勝(1000万)1600m","芝阪神1勝(500万)1600m","芝阪神未勝利1600m","芝阪神新馬1600m","芝阪神G1、G2、G3、L、OP1800m","芝阪神3勝(1600万)1800m","芝阪神2勝(1000万)1800m","芝阪神1勝(500万)1800m","芝阪神未勝利1800m","芝阪神新馬1800m","芝阪神G1、G2、G3、L、OP2000m","芝阪神3勝(1600万)2000m","芝阪神2勝(1000万)2000m","芝阪神1勝(500万)2000m","芝阪神未勝利2000m","芝阪神新馬2000m","芝阪神G1、G2、G3、L、OP2200m","芝阪神3勝(1600万)2200m","芝阪神2勝(1000万)2200m","芝阪神1勝(500万)2200m","芝阪神未勝利2200m","芝阪神G1、G2、G3、L、OP2400m","芝阪神3勝(1600万)2400m","芝阪神2勝(1000万)2400m","芝阪神1勝(500万)2400m","芝阪神未勝利2400m","芝阪神G1、G2、G3、L、OP2600m","芝阪神2勝(1000万)2600m","芝阪神1勝(500万)2600m","芝阪神未勝利2600m","芝阪神G1、G2、G3、L、OP3000m","芝阪神3勝(1600万)3000m","芝阪神G1、G2、G3、L、OP3200m","芝阪神3勝(1600万)3200m","芝小倉G1、G2、G3、L、OP1200m","芝小倉3勝(1600万)1200m","芝小倉2勝(1000万)1200m","芝小倉1勝(500万)1200m","芝小倉未勝利1200m","芝小倉新馬1200m","芝小倉1勝(500万)1700m","芝小倉G1、G2、G3、L、OP1800m","芝小倉3勝(1600万)1800m","芝小倉2勝(1000万)1800m","芝小倉1勝(500万)1800m","芝小倉未勝利1800m","芝小倉新馬1800m","芝小倉G1、G2、G3、L、OP2000m","芝小倉3勝(1600万)2000m","芝小倉2勝(1000万)2000m","芝小倉1勝(500万)2000m","芝小倉未勝利2000m","芝小倉新馬2000m","芝小倉2勝(1000万)2600m","芝小倉1勝(500万)2600m","芝小倉未勝利2600m","ダート札幌2勝(1000万)1000以下m","ダート札幌1勝(500万)1000以下m","ダート札幌未勝利1000以下m","ダート札幌新馬1000以下m","ダート札幌G1、G2、G3、L、OP1700m","ダート札幌3勝(1600万)1700m","ダート札幌2勝(1000万)1700m","ダート札幌1勝(500万)1700m","ダート札幌未勝利1700m","ダート札幌新馬1700m","ダート札幌1勝(500万)2400m","ダート札幌未勝利2400m","ダート函館2勝(1000万)1000以下m","ダート函館1勝(500万)1000以下m","ダート函館未勝利1000以下m","ダート函館新馬1000以下m","ダート函館G1、G2、G3、L、OP1700m","ダート函館3勝(1600万)1700m","ダート函館2勝(1000万)1700m","ダート函館1勝(500万)1700m","ダート函館未勝利1700m","ダート函館新馬1700m","ダート函館1勝(500万)2400m","ダート函館未勝利2400m","ダート福島3勝(1600万)1150m","ダート福島2勝(1000万)1150m","ダート福島1勝(500万)1150m","ダート福島未勝利1150m","ダート福島新馬1150m","ダート福島G1、G2、G3、L、OP1700m","ダート福島3勝(1600万)1700m","ダート福島2勝(1000万)1700m","ダート福島1勝(500万)1700m","ダート福島未勝利1700m","ダート福島新馬1700m","ダート福島1勝(500万)2400m","ダート新潟G1、G2、G3、L、OP1200m","ダート新潟3勝(1600万)1200m","ダート新潟2勝(1000万)1200m","ダート新潟1勝(500万)1200m","ダート新潟未勝利1200m","ダート新潟新馬1200m","ダート新潟G1、G2、G3、L、OP1800m","ダート新潟3勝(1600万)1800m","ダート新潟2勝(1000万)1800m","ダート新潟1勝(500万)1800m","ダート新潟未勝利1800m","ダート新潟新馬1800m","ダート新潟1勝(500万)2500m","ダート新潟未勝利2500m","ダート東京3勝(1600万)1300m","ダート東京2勝(1000万)1300m","ダート東京1勝(500万)1300m","ダート東京未勝利1300m","ダート東京新馬1300m","ダート東京G1、G2、G3、L、OP1400m","ダート東京3勝(1600万)1400m","ダート東京2勝(1000万)1400m","ダート東京1勝(500万)1400m","ダート東京未勝利1400m","ダート東京新馬1400m","ダート東京G1、G2、G3、L、OP1600m","ダート東京3勝(1600万)1600m","ダート東京2勝(1000万)1600m","ダート東京1勝(500万)1600m","ダート東京未勝利1600m","ダート東京新馬1600m","ダート東京G1、G2、G3、L、OP2100m","ダート東京3勝(1600万)2100m","ダート東京2勝(1000万)2100m","ダート東京1勝(500万)2100m","ダート東京未勝利2100m","ダート中山G1、G2、G3、L、OP1200m","ダート中山3勝(1600万)1200m","ダート中山2勝(1000万)1200m","ダート中山1勝(500万)1200m","ダート中山未勝利1200m","ダート中山新馬1200m","ダート中山G1、G2、G3、L、OP1800m","ダート中山3勝(1600万)1800m","ダート中山2勝(1000万)1800m","ダート中山1勝(500万)1800m","ダート中山未勝利1800m","ダート中山新馬1800m","ダート中山2勝(1000万)2400m","ダート中山1勝(500万)2400m","ダート中山1勝(500万)2500m","ダート中京G1、G2、G3、L、OP1200m","ダート中京3勝(1600万)1200m","ダート中京2勝(1000万)1200m","ダート中京1勝(500万)1200m","ダート中京未勝利1200m","ダート中京新馬1200m","ダート中京G1、G2、G3、L、OP1400m","ダート中京3勝(1600万)1400m","ダート中京2勝(1000万)1400m","ダート中京1勝(500万)1400m","ダート中京未勝利1400m","ダート中京新馬1400m","ダート中京G1、G2、G3、L、OP1800m","ダート中京3勝(1600万)1800m","ダート中京2勝(1000万)1800m","ダート中京1勝(500万)1800m","ダート中京未勝利1800m","ダート中京新馬1800m","ダート中京G1、G2、G3、L、OP1900m","ダート中京3勝(1600万)1900m","ダート中京2勝(1000万)1900m","ダート中京1勝(500万)1900m","ダート中京未勝利1900m","ダート京都G1、G2、G3、L、OP1200m","ダート京都3勝(1600万)1200m","ダート京都2勝(1000万)1200m","ダート京都1勝(500万)1200m","ダート京都未勝利1200m","ダート京都新馬1200m","ダート京都G1、G2、G3、L、OP1400m","ダート京都3勝(1600万)1400m","ダート京都2勝(1000万)1400m","ダート京都1勝(500万)1400m","ダート京都未勝利1400m","ダート京都新馬1400m","ダート京都G1、G2、G3、L、OP1800m","ダート京都3勝(1600万)1800m","ダート京都2勝(1000万)1800m","ダート京都1勝(500万)1800m","ダート京都未勝利1800m","ダート京都新馬1800m","ダート京都G1、G2、G3、L、OP1900m","ダート京都3勝(1600万)1900m","ダート京都2勝(1000万)1900m","ダート京都1勝(500万)1900m","ダート京都未勝利1900m","ダート阪神G1、G2、G3、L、OP1200m","ダート阪神3勝(1600万)1200m","ダート阪神2勝(1000万)1200m","ダート阪神1勝(500万)1200m","ダート阪神未勝利1200m","ダート阪神新馬1200m","ダート阪神G1、G2、G3、L、OP1400m","ダート阪神3勝(1600万)1400m","ダート阪神2勝(1000万)1400m","ダート阪神1勝(500万)1400m","ダート阪神未勝利1400m","ダート阪神新馬1400m","ダート阪神G1、G2、G3、L、OP1800m","ダート阪神3勝(1600万)1800m","ダート阪神2勝(1000万)1800m","ダート阪神1勝(500万)1800m","ダート阪神未勝利1800m","ダート阪神新馬1800m","ダート阪神G1、G2、G3、L、OP2000m","ダート阪神3勝(1600万)2000m","ダート阪神2勝(1000万)2000m","ダート阪神1勝(500万)2000m","ダート阪神未勝利2000m","ダート小倉3勝(1600万)1000以下m","ダート小倉2勝(1000万)1000以下m","ダート小倉1勝(500万)1000以下m","ダート小倉未勝利1000以下m","ダート小倉新馬1000以下m","ダート小倉G1、G2、G3、L、OP1700m","ダート小倉3勝(1600万)1700m","ダート小倉2勝(1000万)1700m","ダート小倉1勝(500万)1700m","ダート小倉未勝利1700m","ダート小倉新馬1700m","ダート小倉1勝(500万)2400m"]

            # --- 高速化のためのマッピング定義 ---
            track_map = {'1': '芝', '2': 'ダート'}
            jyo_map = {
                '01': '札幌', '02': '函館', '03': '福島', '04': '新潟', 
                '05': '東京', '06': '中山', '07': '中京', '08': '京都', 
                '09': '阪神', '10': '小倉'
            }
            class_map = {
                "&grade%5B%5D=1&grade%5B%5D=2&grade%5B%5D=3&grade%5B%5D=11&grade%5B%5D=4": "OP",
                "&grade%5B%5D=5": "3勝",
                "&grade%5B%5D=6": "2勝",
                "&grade%5B%5D=7": "1勝",
                "&grade%5B%5D=9": "未勝利",
                "&grade%5B%5D=8": "新馬"
            }

            # --- Excelファイルの準備とヘッダー書き込み ---
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
            
            # 最初にすべてのヘッダーを書き込む
            for year_idx, (p_start, p_end) in enumerate(target_periods):
                header_year_str = f"{p_start}～{p_end}年" if p_start != p_end else f"{p_start}年"
                base_col = year_idx * 3 + 1
                ws.cell(row=1, column=base_col, value=f"{header_year_str} 条件")
                ws.cell(row=1, column=base_col+1, value=f"{header_year_str} サンプル数")
                ws.cell(row=1, column=base_col+2, value=f"{header_year_str} 平均タイム")
            self.save_workbook(wb, file_path)

            # --- データ取得と処理 ---
            page_number = ["","&page=2","&page=3","&page=4","&page=5","&page=6","&page=7","&page=8","&page=9","&page=10","&page=11","&page=12","&page=13","&page=14","&page=15","&page=16","&page=17","&page=18"]
            # 全期間で一度だけリクエストを送信
            tyousa_kukann = f"&start_year={start_year}&start_mon={start_month}&end_year={end_year}&end_mon={end_month}"

            print(f"処理開始: {start_year}年 ～ {end_year}年")
            print(f"集計期間: {agg_years}年分, スライド間隔: {slide_years}年")
            print(f"保存先: {file_path}")
            print(f"馬場状態クエリ: {babajyoutai}")
            print("-" * 20)

            current_write_row = 2
            for x in range(len(siba_dart_number)):
                if not self.is_running: break
                for y in range(len(keibajyou_number)):
                    if not self.is_running: break
                    for z in range(len(kyori)):
                        if not self.is_running: break
                        for a in range(len(grade_sentence)):
                            if not self.is_running: break
                            
                            # --- 存在しないレース条件へのアクセスをスキップ ---
                            k_track = track_map.get(siba_dart_number[x], "")
                            k_jyo = jyo_map.get(keibajyou_number[y], "")
                            k_class_raw = grade_sentence[a]
                            k_class = class_map.get(k_class_raw, "")
                            k_dist = kyori[z]
                            k_dist_search = "1000以下m" if k_dist == "1000以下" else k_dist + "m"

                            condition_exists = any(
                                k_track in cond and k_jyo in cond and k_dist_search in cond and
                                (("OP" in cond) if k_class == "OP" else (k_class in cond))
                                for cond in shutoku_jyouken
                            )
                            if not condition_exists:
                                continue
                            
                            processed_count += 1
                            all_times_with_year = []
                            combined_string = ""

                            # --- ページネーションを考慮して全期間のデータを取得 ---
                            for page in page_number:
                                if not self.is_running: break
                                time.sleep(random.uniform(1, 1.5))

                                encoded_kyori = urllib.parse.quote(kyori[z], encoding='euc-jp')
                                url = (f"https://db.netkeiba.com/?pid=race_list&word=&track%5B%5D={siba_dart_number[x]}"
                                       f"{tyousa_kukann}&jyo%5B%5D={keibajyou_number[y]}{babajyoutai}{grade_sentence[a]}"
                                       f"&kyori_min=&kyori_max=&kyori%5B%5D={encoded_kyori}&sort=date&list=100{page}")

                                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
                                
                                try:
                                    response = session.get(url, headers=headers)
                                    response.raise_for_status()
                                    response.encoding = 'EUC-JP'
                                    soup = BeautifulSoup(response.text, 'html.parser')
                                except requests.RequestException as e:
                                    print(f"URL取得エラー: {e}")
                                    break

                                search_result_box = soup.find('div', class_='search_result_box')
                                if not search_result_box or "見つかりませんでした" in search_result_box.get_text():
                                    break

                                table = soup.find('table', class_='race_table_01')
                                if not table or not re.search(r'\d+:\d+\.\d+', table.get_text()):
                                    break
                                
                                # 初回ページでのみ条件文字列を生成
                                if not combined_string:
                                    matches = re.findall(r'\[(.*?)\]', search_result_box.get_text(strip=True))
                                    if matches:
                                        if len(matches) > 1: del matches[1]
                                        
                                        track_s = ""
                                        jyo_s = ""
                                        baba_s = ""
                                        class_s = ""
                                        dist_s = ""
                                        
                                        tracks_set = {'芝', 'ダート', '障害'}
                                        jyos_set = {'札幌', '函館', '福島', '新潟', '東京', '中山', '中京', '京都', '阪神', '小倉'}
                                        class_list = []
                                        
                                        for m in matches:
                                            if not m: continue
                                            if m in tracks_set: track_s = m
                                            elif m in jyos_set: jyo_s = m
                                            elif all(c in "良稍重不良、" for c in m): baba_s = m
                                            elif 'm' in m and any(c.isdigit() for c in m): dist_s = m
                                            else: class_list.append(m)
                                        
                                        combined_string = f"{track_s}{jyo_s}{dist_s}{''.join(class_list)}"
                                        if baba_s: combined_string += f"、{baba_s}"
                                        
                                        print(f"--- 取得中: {combined_string} ---")

                                rows = table.find_all('tr')
                                if len(rows) < 2: break

                                header_cells = rows[0].find_all(['th', 'td'])
                                time_col_idx = next((i for i, cell in enumerate(header_cells) if 'タイム' in cell.get_text()), -1)
                                date_col_idx = next((i for i, cell in enumerate(header_cells) if '開催日' in cell.get_text()), -1)

                                if time_col_idx == -1 or date_col_idx == -1:
                                    print("列（タイム or 開催日）が見つかりません。")
                                    break

                                for row in rows[1:]:
                                    cols = row.find_all('td')
                                    if len(cols) > max(time_col_idx, date_col_idx):
                                        time_str = cols[time_col_idx].get_text(strip=True)
                                        date_str = cols[date_col_idx].get_text(strip=True)
                                        if time_str and date_str:
                                            try:
                                                year = int(date_str.split('/')[0])
                                                minutes, seconds = (time_str.split(':') + [0])[:2]
                                                total_seconds = int(minutes) * 60 + float(seconds)
                                                all_times_with_year.append((year, total_seconds))
                                            except (ValueError, IndexError):
                                                continue
                                if len(rows) -1 < 100:
                                    break # 最終ページ
                            
                            # --- 取得した全データから期間ごとに集計しExcelに書き込み ---
                            if not all_times_with_year:
                                continue

                            print(f"--- {combined_string} の計算と書き込み開始 ---")
                            for year_idx, (p_start, p_end) in enumerate(target_periods):
                                base_col = year_idx * 3 + 1
                                filtered_times = [t for y, t in all_times_with_year if p_start <= y <= p_end]
                                
                                if filtered_times:
                                    avg_seconds = sum(filtered_times) / len(filtered_times)
                                    avg_minutes = int(avg_seconds // 60)
                                    avg_remainder_seconds = avg_seconds % 60
                                    formatted_avg_time = f"{avg_minutes}:{avg_remainder_seconds:04.1f}"
                                    
                                    ws.cell(row=current_write_row, column=base_col, value=combined_string)
                                    ws.cell(row=current_write_row, column=base_col + 1, value=len(filtered_times))
                                    ws.cell(row=current_write_row, column=base_col + 2, value=formatted_avg_time)
                                else:
                                    # データがない場合も条件名は書き込む
                                    ws.cell(row=current_write_row, column=base_col, value=combined_string)
                                    ws.cell(row=current_write_row, column=base_col + 1, value=0)
                                    ws.cell(row=current_write_row, column=base_col + 2, value="-")
                            
                            current_write_row += 1
                            self.save_workbook(wb, file_path)
            
            if self.is_running:
                end_time = time.time()
                elapsed_time = end_time - start_time
                hours = int(elapsed_time // 3600)
                minutes = int((elapsed_time % 3600) // 60)
                seconds = int(elapsed_time % 60)
                print(f"\n--- 完了レポート ---")
                print(f"処理した検索条件の組み合わせ数: {processed_count}")
                print(f"スクレイピング合計時間: {hours}時間{minutes}分{seconds}秒")
                
                self.align_excel_rows(file_path, file_path)
            else:
                end_time = time.time()
                elapsed_time = end_time - start_time
                hours = int(elapsed_time // 3600)
                minutes = int((elapsed_time % 3600) // 60)
                seconds = int(elapsed_time % 60)
                print(f"\n--- 中断レポート ---")
                print(f"処理した検索条件の組み合わせ数: {processed_count}")
                print(f"スクレイピング経過時間: {hours}時間{minutes}分{seconds}秒")
                self.status_var.set("中断しました")
                self.status_label.configure(foreground="orange")
                print("処理が中断されました。")

        except Exception as e:
            self.status_var.set("データ取得失敗")
            self.status_label.configure(foreground="red")
            print(f"エラーが発生しました: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            self.is_running = False
            self.root.after(0, lambda: self.run_button.config(state="normal"))
            self.root.after(0, lambda: self.stop_button.config(state="disabled"))

if __name__ == "__main__":
    root = tk.Tk()
    app = NetkeibaScraperApp(root)
    root.mainloop()
